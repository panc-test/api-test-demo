"""
处理测试数据文件的工具类
1、创建一个操作excel文件的类，包含构造器，读取和写入文件的方法。
2、使用faker模块生成测试数据。
3、使用openpyxl模块，将测试数据写入excel文件。
4、读取测试数据，保存在列表中

"""
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from faker import Faker
import json
import random


class DoExcel():

    def __init__(self, filepath, sheetname):
        """
        类的构造器，接受外部传参，实例化类的时候需要带上参数
        :param filepath: excel文件路径
        :param sheetname: sheet页名称
        """
        self.filepath = filepath
        self.sheetname = sheetname
        self.fake = Faker("zh_CN")

    def read_excel(self):
        """
        读取excel文件的方法,将读到的数据存放在一个列表
        :return:
        """
        wb = load_workbook(self.filepath)
        ws: Worksheet = wb[self.sheetname]
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            data.append(row)
        return data

    def write_excel(self):
        """
        写入excel文件的方法
        :return:
        """
        wb = Workbook()  # 实例化一个工作簿
        ws = wb.create_sheet(title=self.sheetname, index=0)  # 创建一个sheet页
        ws.append(['case_id', 'method', 'url', 'params', 'expect_val', 'result_val',
                   'result'])  # 往sheet页中写入数据，注意append接受的参数是一个迭代器

        method = "get"
        url = "http://49.233.108.117:3000/api/v1/topics"
        for x in range(2):
            case_id = x + 1
            params_data = {
                "page": 1,
                "tab": random.choice(["ask", "share", "good", "job"]),
                "limit": 1,
                "mdrender": "false"
            }
            # params_data 是一个字典，写入不到excel中，需要先转换成字符串格式,str()是单引号
            params_data = json.dumps(params_data)
            ws.append([case_id, method, url, params_data])

        wb.save(self.filepath)

    def get_random_tab(self, tab_list):
        """
        #根据入参列表，获取随机数方法
        :param tab_list: 主题分类列表
        :return:
        """
        return self.fake.sentence(ext_word_list=tab_list, nb_words=1)


if __name__ == '__main__':
    doexcel = DoExcel(filepath='../data/topicdata.xlsx', sheetname='indextopic')
    doexcel.write_excel()
    all_data = doexcel.read_excel()
    print(all_data)
